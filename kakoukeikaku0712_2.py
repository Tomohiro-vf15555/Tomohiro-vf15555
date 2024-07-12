import openpyxl
import tkinter as tk
from pathlib import Path
from tkinter import filedialog
from tkinter import messagebox
#画面にメッセージを表示する
root = tk.Tk()
root.withdraw()

#2回ループでファイルを選択する
for i in range(2):
    if i == 0:
        message = "ALCデータを選択してください"
    else:
        message = "加工計画を選択してください"
    #メッセージBOXを表示する
    messagebox.showinfo("ファイル選択",message)
    #選択したファイルのpathを取得
    file_path = filedialog.askopenfilename()
    #エラーメッセージ
    if not file_path:
        print("no file select")
    else:
        print("file_path=",file_path)
    #ファイルを取得する
    wb = openpyxl.load_workbook(file_path)
    #それぞれ別の変数に設定する
    if i == 0:
        wb_alc = wb
    else:
        wb_keikaku = wb
    f_name_s = Path(file_path).stem
    print(i,f_name_s)
#wb_alcのsheet_nameリスト
s_names = wb_keikaku.sheetnames
lot_list = []#行データのリスト
keikaku_list = []#1日分のリスト
day_list = []#Sheet毎のリスト
#sheet名からAFかどうか判定する
page = ""#B3セルの変数(新)
page_z = 0
s_date = ""
p_cnt = ""
#Sheetを順番にループ処理
for s_name in s_names:
    if s_name[:2] == "AF":#Sheetを識別して終了処理
        break
    
    ws_itern = wb_keikaku[s_name]
    page = ws_itern.cell(row = 3,column = 2).value
    if page == "PAGE.01" :#前のSheetと日付が違う
        keikaku_list=[]#リストをリセット
        p_cnt = ""
        print("PAGE.01 OK")
    else:
        p_cnt = "(2)"
    print("B3=",page)
    #列変更ループ処理
    for col_num in range(2, 27,12):
        print("列:",col_num)
        #行の下限までループ処理
        for gyo in ws_itern.iter_rows(min_row = 6, max_row = 63, 
                        min_col = col_num, max_col = col_num + 11):
            row_data = []
            check = gyo[2].value
            #print("check:",check)
            if check.startswith(" "):
                print("DATA END")
                break
            #1行のデータから各セルの値を取り出すループ処理
            for retu in range(12):
                #print("retu=",retu)
                gyo_n = gyo[retu].value
                #I列までのセルから半角スペースを削除する
                if retu < 9:
                    #print("処理前",len(gyo_n),gyo_n)
                    s = len(gyo_n)
                    for s in range(s):
                        gyo_n = gyo_n.replace(" ","")
                    
                    #print("処理後",len(gyo_n),gyo_n)
                #day,lot,kisyuの取り出し
                    #indexが3までの場合に処理
                if retu < 4:
                    g_data = gyo_n
                    
                    print("<4:",retu)
                    #1文字目が"("ではない、空文字列ではない
                    if gyo_n[:1] != "(" and gyo_n != "":
                        if retu == 0:
                            day = gyo_n
                        if retu == 1:
                            lot = gyo_n
                        if retu == 3:
                            kisyu = gyo_n

                        print("not None:",gyo_n)
                    #空文字列なら取得してある値をコピーする
                    if gyo_n == "" or gyo_n[:1] == "(":
                        if retu == 0:
                            g_data = day
                            s_date = day
                        if retu == 1:
                            g_data = lot
                        if retu == 3:
                            g_data = kisyu
                        
                        print("None:",g_data)
                    row_data.append(g_data) 
                    #print("g_data=",row_data[0])               
                else:
                    row_data.append(gyo_n)
                    #print("gyo_n",lot_list[4])
                
                no = gyo[2]
            print("row_data",row_data[3])
            keikaku_list.append(row_data)
            print(f"{s_name}, retu:{retu}.lot{keikaku_list[-1][1]}, kisyu{keikaku_list[-1][3]}{gyo[2].row}")
            
            #print(f"{s_name},retu:{retu}.lot{lot_list[2]},kisyu{lot_list[3]}{gyo[2].row}")
    
    
    
    #シートに転記して保存
    ws_new = wb_alc.create_sheet(title = "keikaku" + s_date + p_cnt)
    for keikaku in keikaku_list:
            ws_new.append(keikaku)

wb_alc.save("独自加工計画try0712.xlsx")
print("独自加工計画saveOK")
